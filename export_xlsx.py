# -*- coding: utf-8 -*-
"""
Генерация Excel-файла из данных, введённых в приложении.
Структура листов и формул — та же, что в bar_inventory.py:
  - "single": Наименование | Объём, мл | Кол-во, шт | Остаток, г | ИТОГО, г
  - "multi":  Наименование | Объём1/Кол-во1 | Объём2/Кол-во2 | Объём3/Кол-во3
              | Остаток, г | ИТОГО, г
  - "misc":   Наименование | Ед. изм. | Кол-во
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from data import CATEGORIES, ALT_VOLUMES

HEADER_FILL = PatternFill("solid", fgColor="6B2B2B")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="6B2B2B")
CALC_FONT = Font(color="595959")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _num(value):
    """Строка из поля ввода -> число для Excel, либо None если пусто/некорректно."""
    if value is None:
        return None
    value = str(value).strip().replace(",", ".")
    if not value:
        return None
    try:
        return float(value)
    except ValueError:
        return None


def _style_header(ws, ncols, row):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _build_single_sheet(wb, title, items, values_for_index):
    ws = wb.create_sheet(title[:31])
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    cols = ["№", "Наименование", "Объём бутылки, мл",
            "Кол-во бутылок, шт", "Остаток открытой бутылки, г", "ИТОГО, г"]
    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    header_row = 3
    for i, c in enumerate(cols, start=1):
        ws.cell(row=header_row, column=i, value=c)
    _style_header(ws, len(cols), header_row)

    first_row = header_row + 1
    for idx, (name, volume) in enumerate(items):
        r = first_row + idx
        vals = values_for_index(idx)
        ws.cell(row=r, column=1, value=idx + 1)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=volume)
        ws.cell(row=r, column=4, value=_num(vals.get("qty")))
        ws.cell(row=r, column=5, value=_num(vals.get("rest")))
        total = ws.cell(row=r, column=6, value=f"=C{r}*D{r}+E{r}")
        total.font = CALC_FONT
        for c in range(1, len(cols) + 1):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = LEFT if c == 2 else CENTER

    widths = [5, 44, 14, 16, 20, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{first_row}"


def _build_multi_sheet(wb, title, items, values_for_index):
    ws = wb.create_sheet(title[:31])
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    cols = ["№", "Наименование",
            "Объём 1, мл", "Кол-во, шт",
            "Объём 2, мл", "Кол-во, шт",
            "Объём 3, мл", "Кол-во, шт",
            "Остаток открытой бутылки, г", "ИТОГО, г"]
    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    header_row = 3
    for i, c in enumerate(cols, start=1):
        ws.cell(row=header_row, column=i, value=c)
    _style_header(ws, len(cols), header_row)

    first_row = header_row + 1
    for idx, (name, volume_spec) in enumerate(items):
        r = first_row + idx
        vals = values_for_index(idx)

        if isinstance(volume_spec, (tuple, list)):
            default_volume = volume_spec[0]
            alt = list(volume_spec[1:])
        else:
            default_volume = volume_spec
            alt = [v for v in ALT_VOLUMES if v != default_volume]
        v2 = alt[0] if len(alt) > 0 else ""
        v3 = alt[1] if len(alt) > 1 else ""

        ws.cell(row=r, column=1, value=idx + 1)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=default_volume)
        ws.cell(row=r, column=4, value=_num(vals.get("qty1")))
        ws.cell(row=r, column=5, value=v2)
        ws.cell(row=r, column=6, value=_num(vals.get("qty2")))
        ws.cell(row=r, column=7, value=v3)
        ws.cell(row=r, column=8, value=_num(vals.get("qty3")))
        ws.cell(row=r, column=9, value=_num(vals.get("rest")))
        total = ws.cell(row=r, column=10,
                         value=f"=C{r}*D{r}+E{r}*F{r}+G{r}*H{r}+I{r}")
        total.font = CALC_FONT
        for c in range(1, len(cols) + 1):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = LEFT if c == 2 else CENTER

    widths = [5, 34, 10, 8, 10, 8, 10, 8, 20, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{first_row}"


def _build_misc_sheet(wb, title, items, values_for_index):
    ws = wb.create_sheet(title[:31])
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    cols = ["№", "Наименование", "Ед. изм.", "Кол-во"]
    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    header_row = 3
    for i, c in enumerate(cols, start=1):
        ws.cell(row=header_row, column=i, value=c)
    _style_header(ws, len(cols), header_row)

    first_row = header_row + 1
    for idx, (name, unit) in enumerate(items):
        r = first_row + idx
        vals = values_for_index(idx)
        ws.cell(row=r, column=1, value=idx + 1)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=unit)
        ws.cell(row=r, column=4, value=_num(vals.get("qty")))
        for c in range(1, len(cols) + 1):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = LEFT if c == 2 else CENTER

    widths = [5, 38, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{first_row}"


def build_workbook(store):
    """store — словарь из storage.load_store()."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for cat in CATEGORIES:
        cat_store = store.get(cat["key"], {})

        def values_for_index(idx, _cat_store=cat_store):
            return _cat_store.get(str(idx), {})

        if cat["type"] == "single":
            _build_single_sheet(wb, cat["title"], cat["items"], values_for_index)
        elif cat["type"] == "multi":
            _build_multi_sheet(wb, cat["title"], cat["items"], values_for_index)
        else:
            _build_misc_sheet(wb, cat["title"], cat["items"], values_for_index)

    return wb


def export_to_file(store, filepath):
    wb = build_workbook(store)
    wb.save(filepath)
    return filepath
